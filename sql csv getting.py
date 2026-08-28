import csv
import io
import sqlite3
import urllib.request
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

#########################################################
## live portfolio import
#########################################################
DB_PATH = Path.home() / "Library/Application Support/Ledger/ledger.db"
OUT_DIR = Path.home() / "Downloads"
LOOKBACK_YEARS = 3
VAR_HORIZONS = (1, 5, 30)
VAR_CONFIDENCE = 0.95
TRADING_DAYS = 252  # Excel/textbook Sharpe. Ledger's in-app Sharpe uses 365.
FRED_DGS3MO = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=DGS3MO"


def trailing_rf(default: float = 0.05) -> float:
    """Latest 3-month T-bill yield from FRED, as a decimal."""
    try:
        with urllib.request.urlopen(FRED_DGS3MO, timeout=10) as resp:
            text = resp.read().decode()
        latest = None
        for row in csv.DictReader(io.StringIO(text)):
            raw = row.get("DGS3MO", ".")
            if raw and raw != ".":
                latest = float(raw)
        return latest / 100.0 if latest is not None else default
    except Exception:
        return default


conn = sqlite3.connect(str(DB_PATH))

account_value = conn.execute("""
    SELECT SUM(a.current_balance)
    FROM accounts a
    JOIN items i ON i.id = a.item_id
    WHERE a.type = 'investment'
      AND i.item_id NOT IN ('manual_import', 'test_item', 'crypto_wallet')
""").fetchone()[0]

holdings = pd.read_sql_query("""
    SELECT s.ticker_symbol AS ticker, s.type AS security_type,
    SUM(h.institution_value) AS value
    FROM holdings h
    JOIN securities s ON s.id = h.security_id
    JOIN accounts a ON a.id = h.account_id
    JOIN items i ON i.id = a.item_id
    WHERE a.type = 'investment'
      AND s.ticker_symbol IS NOT NULL
      AND IFNULL(s.is_cash_equivalent, 0) = 0
      AND i.item_id NOT IN ('manual_import', 'test_item', 'crypto_wallet')
    GROUP BY s.ticker_symbol
""", conn)
values = holdings.set_index("ticker")["value"]
weights = values / values.sum()

tickers = list(weights.index)
placeholders = ",".join("?" * len(tickers))
sql_price = pd.read_sql_query(
    f"SELECT ticker, price_date, close_price FROM market_prices WHERE ticker IN ({placeholders})",
    conn,
    params=tickers,
)
spy = pd.read_sql_query(
    "SELECT price_date, close_price FROM market_prices WHERE ticker = 'SPY'",
    conn,
)
conn.close()
#########################################################
## prices → daily portfolio returns (what Excel needs)
#########################################################
rf_rate = trailing_rf()

sql_price["price_date"] = pd.to_datetime(sql_price["price_date"])
spy["price_date"] = pd.to_datetime(spy["price_date"])
sql_price = sql_price.sort_values("price_date")
spy = spy.sort_values("price_date")

close_df = sql_price.pivot(index="price_date", columns="ticker", values="close_price")
close_df = close_df[close_df.columns.intersection(weights.index)]
end = close_df.index.max()
start = end - pd.DateOffset(years=LOOKBACK_YEARS)
close_df = close_df.loc[start:end]
min_obs = max(252, int(0.95 * len(close_df)))
close_df = close_df.loc[:, close_df.count() >= min_obs]
close_df = close_df.dropna()

weights = weights.reindex(close_df.columns).dropna()
weights = weights / weights.sum()

log_returns = np.log(close_df / close_df.shift(1))
port_log = log_returns.mul(weights, axis=1).sum(axis=1, min_count=1).dropna()
port_simple = np.exp(port_log) - 1

ret_5d = np.exp(port_log.rolling(5).sum()) - 1
ret_30d = np.exp(port_log.rolling(30).sum()) - 1

#########################################################
## one .xlsx — Results has the formulas, Daily is the raw series
#########################################################
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
input_fill = PatternFill("solid", fgColor="FFF2CC")
section_font = Font(bold=True, size=14)
thin = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header(ws, row=1, cols=None):
    last_col = cols or ws.max_column
    for col in range(1, last_col + 1):
        cell = ws.cell(row, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


wb = Workbook()

# ── Results ──────────────────────────────────────────────
ws = wb.active
ws.title = "Results"
last = len(port_simple) + 1  # Daily header is row 1, data starts row 2

ws["A1"] = "Sharpe & VaR"
ws["A1"].font = section_font
ws.merge_cells("A1:C1")
ws["A2"] = (
    "Yellow cells are inputs. Column B on this sheet is live Excel formulas "
    "against the Daily tab — change a yellow cell and the numbers recalc."
)
ws.merge_cells("A2:C2")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[2].height = 36

ws["A4"] = "Inputs"
ws["A4"].font = Font(bold=True)
for row, (label, value, fmt, note) in enumerate((
    ("Portfolio value", float(account_value), '"$"#,##0.00', "Net investment accounts"),
    ("Risk-free rate", rf_rate, "0.00%", "3-month T-bill (FRED DGS3MO)"),
    ("Trading days", TRADING_DAYS, "0", "252 = Excel/textbook. Ledger app uses 365."),
    ("VaR confidence", VAR_CONFIDENCE, "0%", "95% → 5th percentile of history"),
), start=5):
    ws.cell(row, 1, label)
    cell = ws.cell(row, 2, value)
    cell.number_format = fmt
    cell.fill = input_fill
    cell.border = thin
    ws.cell(row, 3, note)

ws["A10"] = "Results"
ws["A10"].font = Font(bold=True)
ws["C10"] = "What this is"

results_rows = [
    (11, "Sharpe ratio",
     f"=(AVERAGE(Daily!B2:B{last})*$B$7-$B$6)/(STDEV.S(Daily!B2:B{last})*SQRT($B$7))",
     "0.00",
     "Extra annual return per unit of volatility. ~1 is fine, >1 is strong."),
    (12, "1-day 95% VaR",
     f"=-PERCENTILE.INC(Daily!C2:C{last},1-$B$8)",
     '"$"#,##0.00',
     "Dollar loss that was only worse on 5% of days."),
    (13, "5-day 95% VaR",
     f"=-PERCENTILE.INC(Daily!E2:E{last},1-$B$8)",
     '"$"#,##0.00',
     "Same idea over a rolling 5-day window. Blanks at the top are ignored."),
    (14, "30-day 95% VaR",
     f"=-PERCENTILE.INC(Daily!G2:G{last},1-$B$8)",
     '"$"#,##0.00',
     "Same idea over a rolling 30-day window."),
    (15, "1-day 95% VaR %",
     f"=-PERCENTILE.INC(Daily!B2:B{last},1-$B$8)",
     "0.00%",
     "Same as 1-day VaR, as a % of portfolio value."),
]
for row, label, formula, fmt, note in results_rows:
    ws.cell(row, 1, label)
    cell = ws.cell(row, 2, formula)
    cell.number_format = fmt
    cell.font = Font(bold=True)
    ws.cell(row, 3, note)

ws["A17"] = "How to use Daily if you want to type the formulas yourself"
ws["A17"].font = Font(bold=True)
ws.merge_cells("A18:C21")
ws["A18"] = (
    "Daily!B = one-day portfolio return (percent).\n"
    "Daily!C = that return × portfolio value (dollars).\n"
    "Sharpe = (AVERAGE(B)×trading days − risk-free) / (STDEV.S(B)×SQRT(trading days))\n"
    "VaR $ = −PERCENTILE.INC(C, 5%)     VaR % = −PERCENTILE.INC(B, 5%)"
)
ws["A18"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[18].height = 72

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 62

# ── Daily ────────────────────────────────────────────────
ws_d = wb.create_sheet("Daily")
ws_d.append([
    "Date",
    "Daily return",
    "1-day P&L ($)",
    "5-day return",
    "5-day P&L ($)",
    "30-day return",
    "30-day P&L ($)",
])
for dt, r1, r5, r30 in zip(port_simple.index, port_simple, ret_5d, ret_30d):
    ws_d.append([
        dt.date() if hasattr(dt, "date") else dt,
        float(r1),
        float(r1) * float(account_value),
        None if pd.isna(r5) else float(r5),
        None if pd.isna(r5) else float(r5) * float(account_value),
        None if pd.isna(r30) else float(r30),
        None if pd.isna(r30) else float(r30) * float(account_value),
    ])
style_header(ws_d)
for row in ws_d.iter_rows(min_row=2, max_row=ws_d.max_row):
    row[0].number_format = "YYYY-MM-DD"
    row[1].number_format = "0.00%"
    row[2].number_format = '"$"#,##0.00'
    row[3].number_format = "0.00%"
    row[4].number_format = '"$"#,##0.00'
    row[5].number_format = "0.00%"
    row[6].number_format = '"$"#,##0.00'
for i, width in enumerate((12, 14, 16, 14, 16, 15, 16), start=1):
    ws_d.column_dimensions[get_column_letter(i)].width = width

# ── Weights ──────────────────────────────────────────────
ws_w = wb.create_sheet("Weights")
ws_w.append(["Ticker", "Weight", "Value ($)"])
for ticker, weight in weights.items():
    ws_w.append([ticker, float(weight), float(values.reindex([ticker]).iloc[0])])
style_header(ws_w)
for row in ws_w.iter_rows(min_row=2, max_row=ws_w.max_row):
    row[1].number_format = "0.00%"
    row[2].number_format = '"$"#,##0.00'
ws_w.column_dimensions["A"].width = 12
ws_w.column_dimensions["B"].width = 12
ws_w.column_dimensions["C"].width = 16

# ── Prices ───────────────────────────────────────────────
ws_p = wb.create_sheet("Prices")
price_frame = close_df.copy()
price_frame.index = [d.date() if hasattr(d, "date") else d for d in price_frame.index]
price_frame.index.name = "Date"
for i, row in enumerate(dataframe_to_rows(price_frame, index=True, header=True)):
    if i == 1:  # skip openpyxl's extra integer header row
        continue
    ws_p.append(row)
style_header(ws_p)
for row in ws_p.iter_rows(min_row=2, max_row=ws_p.max_row):
    row[0].number_format = "YYYY-MM-DD"
    for cell in row[1:]:
        if cell.value is not None:
            cell.number_format = '"$"#,##0.00'
ws_p.column_dimensions["A"].width = 12
for col in range(2, ws_p.max_column + 1):
    ws_p.column_dimensions[get_column_letter(col)].width = 11

OUT_DIR.mkdir(parents=True, exist_ok=True)
out_path = OUT_DIR / "portfolio_var_sharpe.xlsx"
wb.save(out_path)

for old in OUT_DIR.glob("excel_var_sharpe_*.csv"):
    old.unlink()

print(f"wrote {out_path}")
print(f"  {len(port_simple)} days, {len(weights)} tickers, value={float(account_value):,.2f}, rf={rf_rate:.4%}")
